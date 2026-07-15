import openpyxl
from copy import copy

SRC = r"C:\Users\crist\Projetos\A2A - Projato Econosystens 2.0\mensagens e fotos debug\exportedPrice.xlsx"
DST = r"C:\Users\crist\Projetos\A2A - Projato Econosystens 2.0\mensagens e fotos debug\exportedPrice_fixed.xlsx"

PLANS = [
    ("Starter", "starter", 99700, 166.17),
    ("Professional", "professional", 239100, 398.50),
    ("Enterprise", "enterprise", 468500, 780.83),
    ("Full Suite", "full_suite", 949700, 1582.83),
    ("Compliance Pack", "compliance_pack", 239100, 398.50),
    ("Regulatory Starter", "regulatory_starter", 59000, 98.33),
    ("Regulatory Professional", "regulatory_professional", 149000, 248.33),
    ("Regulatory Full", "regulatory_full", 349000, 581.67),
    ("ESG + Carbono", "esg_carbon_pack", 249000, 415.00),
    ("Microsoft Pack", "microsoft_pack", 448200, 747.00),
]

wb = openpyxl.load_workbook(SRC)

countries_enabled = [
    ("BR", "Brazil", "Yes", "BRL"),
    ("US", "United States", "Yes", "USD"),
    ("PT", "Portugal", "Yes", "EUR"),
]

for plan_title, plan_id, brl_cents, usd_price in PLANS:
    ws = wb.create_sheet(title=plan_id[:31])
    ws.cell(row=1, column=1, value="PublisherId")
    ws.cell(row=1, column=2, value="projeto-engenharia")
    ws.cell(row=2, column=1, value="OfferId")
    ws.cell(row=2, column=2, value="ecosystem-aec-regulatory")
    ws.cell(row=3, column=1, value="OfferTitle")
    ws.cell(row=3, column=2, value="EcoSystem AEC + Regulatory")
    ws.cell(row=4, column=1, value="PlanId")
    ws.cell(row=4, column=2, value=plan_id)
    ws.cell(row=5, column=1, value="PlanTitle")
    ws.cell(row=5, column=2, value=plan_title)
    ws.cell(row=6, column=1, value="Pricing Category")
    ws.cell(row=6, column=2, value="Standard")
    ws.cell(row=7, column=1, value="Is Simplified Currency Pricing?")
    ws.cell(row=7, column=2, value="Yes")
    ws.cell(row=8, column=1, value="Pricing Model")
    ws.cell(row=8, column=2, value="Site")
    ws.cell(row=9, column=1, value="Country/Region Code")
    ws.cell(row=9, column=2, value="Country/Region Name")
    ws.cell(row=9, column=3, value="Tax Remit Status")
    ws.cell(row=9, column=4, value="Currency")
    ws.cell(row=9, column=5, value="Enabled")
    ws.cell(row=9, column=6, value="One-time PMT")

    row = 10
    for code, name, tax, currency in countries_enabled:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=tax)
        ws.cell(row=row, column=4, value=currency)
        ws.cell(row=row, column=5, value="Yes")
        if currency == "BRL":
            ws.cell(row=row, column=6, value=brl_cents / 100)
        elif currency == "USD":
            ws.cell(row=row, column=6, value=usd_price)
        elif currency == "EUR":
            ws.cell(row=row, column=6, value=round(usd_price * 0.92, 2))
        row += 1

wb.save(DST)
print(f"Planilha corrigida salva em: {DST}")
print(f"\nPlanos incluidos:")
for plan_title, plan_id, brl_cents, usd_price in PLANS:
    print(f"  {plan_title} ({plan_id}): R$ {brl_cents/100:.2f} / USD$ {usd_price:.2f}")
print(f"\nPaises ativos: Brasil (BRL), EUA (USD), Portugal (EUR)")
print("\nInstrucoes:")
print("1. Abra o arquivo no Excel")
print("2. No Partner Center, va em Pricing")
print("3. Importe a planilha corrigida")
