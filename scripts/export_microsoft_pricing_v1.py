import openpyxl

DST = r"C:\Users\crist\Projetos\A2A - Projato Econosystens 2.0\mensagens e fotos debug\microsoft_pricing_v1.xlsx"

PLANS = [
    ("Starter - Spec Analyst", "starter", 997, 166.17),
    ("Professional - 3 Agentes", "professional", 2391, 398.50),
    ("Enterprise - 5 Agentes", "enterprise", 4685, 780.83),
    ("Full Suite - 21 Agentes", "full_suite", 9497, 1582.83),
    ("Compliance Pack", "compliance_pack", 2391, 398.50),
    ("Regulatory Starter", "regulatory_starter", 590, 98.33),
    ("Regulatory Professional", "regulatory_professional", 1490, 248.33),
    ("Regulatory Full", "regulatory_full", 3490, 581.67),
    ("ESG + Carbono", "esg_carbon_pack", 2490, 415.00),
    ("Microsoft Pack", "microsoft_pack", 4482, 747.00),
]

COUNTRIES = [
    ("BR", "Brazil", "Yes", "BRL"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

for plan_title, plan_id, brl_price, usd_price in PLANS:
    ws = wb.create_sheet(title=plan_id[:31])
    data = [
        ("PublisherId", "projeto-engenharia"),
        ("OfferId", "ecosystem-aec-regulatory"),
        ("OfferTitle", "EcoSystem AEC + Regulatory"),
        ("PlanId", plan_id),
        ("PlanTitle", plan_title),
        ("Pricing Category", "Standard"),
        ("Is Simplified Currency Pricing?", "Yes"),
        ("Pricing Model", "Site"),
        ("Country/Region Code", "Country/Region Name", "Tax Remit Status", "Currency", "Enabled", "Monthly Price"),
    ]
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    row = 10
    for code, name, tax, currency in COUNTRIES:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=tax)
        ws.cell(row=row, column=4, value=currency)
        ws.cell(row=row, column=5, value="Yes")
        ws.cell(row=row, column=6, value=brl_price)
        row += 1

wb.save(DST)
print(f"Planilha salva em: {DST}")
print(f"\n10 planos incluidos (apenas Brasil ativo):")
for plan_title, plan_id, brl_price, usd_price in PLANS:
    print(f"  {plan_id}: {plan_title} - R$ {brl_price:.2f}/mes")
print("\nInstrucoes:")
print("1. Acesse Partner Center")
print("2. Oferta: ecosystem-aec-regulatory")
print("3. Va em Pricing -> Import")
print("4. Selecione o arquivo gerado")
print("5. Salve e publique")
